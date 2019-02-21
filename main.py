import tkinter as tk
from tkinter import messagebox
from tkinter import *
from openpyxl import Workbook
from datetime import datetime


class MyApp:
    list_ttn_for_screen = ''    # Used for showing ttn list on screen
    list_ttn = []               # Used for saving ttn list

    def __init__(self,  master):
        self.master = master


        '''''''''''''''''screen section'''''''''''''''''
        self.screen = Label(bg='#ccc')
        self.screen.place(relx=.5, rely=.1, anchor="c", height=250, width=800)
        self.screen['text'] = 'Список накладных.'


        '''''''''''''''''ttn section'''''''''''''''''
        self.ttn_name = StringVar()
        self.ttn_date = StringVar()
        self.ttn_sum = StringVar()

        self.ttn_name_entry = Entry(textvariable=self.ttn_name)
        self.ttn_date_entry = Entry(textvariable=self.ttn_date)
        self.ttn_sum_entry = Entry(textvariable=self.ttn_sum)

        self.ttn_name_entry.place(relx=.1, rely=.4, anchor="c")
        self.ttn_date_entry.place(relx=.3, rely=.4, anchor="c")
        self.ttn_sum_entry.place(relx=.5, rely=.4, anchor="c")

        self.message_button = Button(text="Добавить накладную", command=self.ttn_list)
        self.message_button.place(relx=.7, rely=.4, anchor="c")


        '''''''''''''''''ttn section'''''''''''''''''

        self.file_name = StringVar()
        self.file_name_entry = Entry(textvariable=self.file_name)
        self.file_name_entry.place(relx=.5, rely=.6, anchor="c")

        self.message_button = Button(text="Россчитать", command=self.create_document_exel)
        self.message_button.place(relx=.7, rely=.6, anchor="c")

        self.message_button = Button(text="TEST", command=self.str_sum_of_debt)
        self.message_button.place(relx=.7, rely=.8, anchor="c")

    def ttn_list(self):
        a = self.ttn_name.get()
        b = self.ttn_date.get()
        c = self.ttn_sum.get()
        self.list_ttn.append([a,b,c])
        self.list_ttn_for_screen = self.list_ttn_for_screen +('ТТН ' + a + ' від ' +  b + ' на загальну суму ' + c + ' грн. \n')
        self.screen_out(self.list_ttn_for_screen)

    def screen_out(self, x):
        self.screen['text'] = x

    def create_document_exel(self):
        wb = Workbook()
        ws = wb.active
        ws['B2'] = 'РОЗРАХУНОК ЗАБОРГОВАНОСТІ ТА ШТРАФНИХ САНКЦІЙ'
        ws['A3'] = 'Дебіторська заборгованість Позивача розраховується, як сума вартості тоовару поставленого за всіми неоплаченими ТТН'
        ws.append([str(self.str_sum_of_debt()) + str(self.sum_of_debt())])
        ws.append([self.sum_of_debt()])
        wb.save(str(self.file_name.get()) + ".xlsx")

    def sum_of_debt(self):
        x = 0
        for i in self.list_ttn:
            x = x + round(float(i[2]), 2)
        print(x)
        return x

    def str_sum_of_debt(self):
        x = ''
        for i in self.list_ttn:
            x = x + ' + ' + i[2]
        x = x + ' = '
        x = x[2::]
        print(x)
        return x


    # def get_first_date(self):
    #     return self.first_date.get()
    #
    # def get_last_date(self):
    #     return self.last_date.get()
    #
    #
    # def days_between(self):
    #     ''''''''''''''''return the difference between the dates in days'''''''''''''''
    #     d1 = datetime.strptime(self.get_first_date(), "%Y-%m-%d")
    #     d2 = datetime.strptime(self.get_last_date(), "%Y-%m-%d")
    #     print((d1 - d2).days)



def main():
    root = tk.Tk()
    root.geometry('850x550+250+100')
    app = MyApp(root)
    root.mainloop()

if __name__ == '__main__':
    main()

